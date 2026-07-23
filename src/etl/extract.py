"""
Extração dos dados brutos da planilha "Jogos Zerados".

A aba principal ("Jogos Zerados") tem, além da tabela de jogos (colunas A a J),
outras mini-tabelas soltas nas colunas mais à direita (ranking de notas,
desafios, etc). Este módulo lê SOMENTE a tabela de jogos, sem tentar
interpretar ou corrigir nada — a limpeza fica isolada em `clean.py`.
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any

import re
import zipfile
from xml.etree import ElementTree

import openpyxl

# Mapeamento explícito das colunas usadas na aba "Jogos Zerados".
# Se a estrutura da planilha mudar, é só ajustar aqui.
COLUNAS = {
    "ordem": 1,
    "nome": 2,
    "console": 3,
    "genero": 4,
    "tipo": 5,
    "data": 6,
    "tempo": 7,
    "nota": 8,
    "dificuldade": 9,
    "condicao_zeramento": 10,
}

ABA_PRINCIPAL = "Jogos Zerados"
LINHA_INICIAL = 2  # linha 1 é cabeçalho

ABA_DROPADOS = "Jogos Dropados"
LINHA_INICIAL_DROPADOS = 6  # linha 5 é cabeçalho ("NOTA", "JOGOS QUE EU DROPEI", "CONSOLE")

ABA_DESAFIOS = "Desafios"


@dataclass
class JogoBruto:
    """Representa uma linha da planilha, sem nenhum tratamento."""

    linha_planilha: int
    ordem: Any
    nome: Any
    console: Any
    genero: Any
    tipo: Any
    data: Any
    tempo: Any
    nota: Any
    dificuldade: Any
    condicao_zeramento: Any
    comentario_pessoal: Any = None


@dataclass
class JogoDropadoBruto:
    linha_planilha: int
    nota: Any
    nome: Any
    console: Any


@dataclass
class DesafioBruto:
    ano: int
    progresso: Any
    descricao: Any


def extrair_dropados(caminho_planilha: str | Path) -> list[JogoDropadoBruto]:
    """Lê a aba 'Jogos Dropados' (colunas C=nota, D=nome, G=console,
    cabeçalho na linha 5)."""
    wb = openpyxl.load_workbook(caminho_planilha, data_only=True)
    ws = wb[ABA_DROPADOS]

    registros: list[JogoDropadoBruto] = []
    linhas_vazias_seguidas = 0
    linha = LINHA_INICIAL_DROPADOS

    while linhas_vazias_seguidas < 3:
        nome = ws.cell(row=linha, column=4).value  # coluna D

        if nome is None:
            linhas_vazias_seguidas += 1
            linha += 1
            continue

        linhas_vazias_seguidas = 0
        registros.append(
            JogoDropadoBruto(
                linha_planilha=linha,
                nota=ws.cell(row=linha, column=3).value,  # coluna C
                nome=nome,
                console=ws.cell(row=linha, column=7).value,  # coluna G
            )
        )
        linha += 1

    return registros


def extrair_desafios(caminho_planilha: str | Path) -> list[DesafioBruto]:
    """Lê a aba 'Desafios'.

    Essa aba não tem uma tabela regular: tem cabeçalhos de ano soltos
    (ex: 2025.0 sozinho na coluna B) seguidos de linhas de desafio
    (progresso 0-1 na coluna B + descrição na coluna C). Também existem
    linhas "placeholder" (progresso sem descrição) que representam
    metas ainda não definidas — essas são ignoradas.
    """
    wb = openpyxl.load_workbook(caminho_planilha, data_only=True)
    ws = wb[ABA_DESAFIOS]

    registros: list[DesafioBruto] = []
    ano_atual: int | None = None

    for row in ws.iter_rows(min_row=1, max_row=100, max_col=3):
        valor_b = row[1].value  # coluna B
        valor_c = row[2].value  # coluna C

        if isinstance(valor_b, (int, float)) and valor_b >= 2000:
            ano_atual = int(valor_b)
            continue

        if valor_c is None or ano_atual is None:
            continue  # nota solta, placeholder vazio, ou fora de uma seção de ano

        registros.append(
            DesafioBruto(ano=ano_atual, progresso=valor_b, descricao=valor_c)
        )

    return registros



NS_THREADED_COMMENT = "{http://schemas.microsoft.com/office/spreadsheetml/2018/threadedcomments}"
COLUNA_COMENTARIOS = "H"  # comentários pessoais foram anexados às células de nota


def extrair_comentarios_pessoais(caminho_planilha: str | Path) -> dict[int, str]:
    """Extrai os comentários encadeados ("threaded comments") do Excel.

    Achado importante: o autor da planilha não escrevia suas opiniões
    numa coluna — ele usava o recurso de comentário do Excel (clique
    direito > novo comentário) em cima da célula de nota. O openpyxl
    não expõe o texto desses comentários através de `cell.comment`
    (só mostra um aviso de compatibilidade), então lemos o XML
    interno do .xlsx diretamente.

    Retorna um dicionário {linha: texto_completo}, já juntando
    comentário original + respostas em sequência (uma pessoa às vezes
    complementa a própria opinião com uma segunda mensagem, tipo um
    "Obs:" adicional).
    """
    comentarios_por_celula: dict[str, list[tuple[str, str]]] = {}

    with zipfile.ZipFile(caminho_planilha) as z:
        arquivos = [n for n in z.namelist() if "threadedComments" in n and n.endswith(".xml")]
        for nome_arquivo in arquivos:
            with z.open(nome_arquivo) as f:
                root = ElementTree.parse(f).getroot()
                for tc in root.findall(f"{NS_THREADED_COMMENT}threadedComment"):
                    ref = tc.get("ref", "")
                    data_hora = tc.get("dT", "")
                    texto_el = tc.find(f"{NS_THREADED_COMMENT}text")
                    texto = texto_el.text if texto_el is not None else ""
                    comentarios_por_celula.setdefault(ref, []).append((data_hora, texto or ""))

    resultado: dict[int, str] = {}
    for ref, itens in comentarios_por_celula.items():
        match = re.match(r"([A-Z]+)(\d+)", ref)
        if not match:
            continue
        coluna, linha = match.groups()
        if coluna != COLUNA_COMENTARIOS:
            continue
        itens.sort(key=lambda item: item[0])  # ordena por data/hora do comentário
        resultado[int(linha)] = "\n\n".join(texto for _, texto in itens if texto)

    return resultado


def extrair_jogos(caminho_planilha: str | Path) -> list[JogoBruto]:
    """Lê a aba 'Jogos Zerados' e retorna uma lista de registros brutos.

    Para de ler quando encontra 3 linhas seguidas sem nome de jogo —
    isso evita varrer as ~50 mil linhas vazias que a planilha reserva
    por causa da fórmula de contagem automática na coluna A.
    """
    wb = openpyxl.load_workbook(caminho_planilha, data_only=True)
    ws = wb[ABA_PRINCIPAL]
    comentarios = extrair_comentarios_pessoais(caminho_planilha)

    registros: list[JogoBruto] = []
    linhas_vazias_seguidas = 0
    linha = LINHA_INICIAL

    while linhas_vazias_seguidas < 3:
        valores = {
            campo: ws.cell(row=linha, column=col).value
            for campo, col in COLUNAS.items()
        }

        if valores["nome"] is None:
            linhas_vazias_seguidas += 1
            linha += 1
            continue

        linhas_vazias_seguidas = 0
        registros.append(
            JogoBruto(
                linha_planilha=linha,
                comentario_pessoal=comentarios.get(linha),
                **valores,
            )
        )
        linha += 1

    return registros
